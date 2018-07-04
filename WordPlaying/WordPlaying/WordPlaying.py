
import openpyxl
#from openpyxl import Workbook
#from openpyxl import load_workbook
#import pyexcel as pe

from time import sleep
import nt
from re import split 

import win32com.client as win32
import sys
import string
import OpenOPC
import os

RANGE = range(3, 8)
""" 
def word():
    word = win32.gencache.EnsureDispatch('Word.Application')
    doc = word.Documents.Add()
    word.Visible = True
    sleep(1)
 
    rng = doc.Range(0,0)
    rng.InsertAfter('Hacking Word with Python\r\n\r\n')
    sleep(1)
    for i in RANGE:
        rng.InsertAfter('Line %d\r\n' % i)
        sleep(1)
    rng.InsertAfter("\r\nPython rules!\r\n")
 
    doc.Close(False)
    word.Application.Quit()
 
if __name__ == '__main__':
    word()
"""

class ClNameConversion:
    'The file is used to convert Gloval Var names to Tab names in Rail Monitor'

    def __init__(self, fileName = 'FileNameTBD.txt'):
        
        
        self.convTable = []
        if (fileName != 'FileNameTBD.txt'):
            with open(fileName,'r',encoding="ansi") as self.fileHnd: #open conversion  file 
                words = ['']
                for line in self.fileHnd:
                    words = line.split('=')
                    if (len(words) > 1):
                        words[0] = words[0].strip()
                        words[1] = words[1].strip()
                        oneRow = {}
                        oneRow['varname'] = words[0]
                        oneRow['tabname'] = words[1].strip('\n')
                        self.convTable.append(oneRow)
            if not(self.fileHnd.closed):
                self.fileHnd.close()
       
    def searchVars(self,varname):
        tabname = ''
        for row in self.convTable:
            if row['varname'] == varname:
                tabname = row['tabname']
                return tabname
        return varname



       



class ClXmlFile:
    'Xml file used in the Rail Monitor to define the sheet for monitoring/changing PLC variables'

    def __init__(self, fileName = 'FileNameTBD.xml', varListFileName = 'nil', LogResourceString = '<LogResource Name="xxxxxxxx" Bild="pictures\\trampng.png\">\n'):

        self.IdNo = 0
        self.varListFileName = varListFileName
        self.xmlFileHnd = open(fileName,'w',encoding="utf8")
        self.xmlFileHnd.write('<?xml version="1.0" encoding="UTF-8"?>\n') #prepare the header of the RailMonitor conf xml file
        self.xmlFileHnd.write('<RailMonitor>\n')
        self.xmlFileHnd.write('  <Settings IP="192.168.0.243" Varlist="..\..\ATLG\VARLIST.csv" Task="500" Timeout="1000"></Settings>\n')
        #self.xmlFileHnd.write('  <Settings IP="192.168.0.4" Varlist="..\..\ATLG\VARLIST.csv" Task="500" Timeout="1000"></Settings>\n')
        self.xmlFileHnd.write(LogResourceString)
    
    def writeLine(self, variableName = 'VariableNameTBD', variableText = 'VariableTextTBD', varListFileName = 'nil', subwordIndex = 0):
        variableNameFound = 0
        variableNameFoundAndWritten = 0
        if (varListFileName == 'nil'):
            varListFileName = self.varListFileName

        if not(self.xmlFileHnd.closed):
            if (varListFileName == 'nil'):
                self.xmlFileHnd.write('    <LogItem ID="" Titel="'  + variableText + '" Name="' + variableName + '" POU="" Einheit="" Faktor="1" Offset="0"></LogItem>\n')
            else:
                #self.varListFileHnd = open(varListFileName,'r',encoding="ansi")
                with open(varListFileName,'r',encoding="ansi") as self.varListFileHnd: #open VARLIST.CSV file 
                    words = ['']                
                    for varListLine in self.varListFileHnd:                       
                        if (-1 != varListLine.find(variableName)): #search for the variable in the VARLIST.CSV
                            variableNameFound += 1
                            words = varListLine.split(";") #split the row of the VARLIST.CSV file by the ;
                            
                            if (len(words) > 2):
                                subword = words[1].split(".")
                                indexSubword = 0
                                if len(subword) >= 2:
                                    if (subword[1] == variableName):
                                        indexSubword = 1
                                    if len(subword) >=3:
                                        if (subword[2] == variableName):
                                            indexSubword = 2
                                else:
                                    indexSubword = 0


                                if (((subword[indexSubword] == variableName) or (subword[0].find(variableName)) ) and (words[2].startswith("#") or words[2].startswith("%")) and not(words[1].startswith("FB_")) and not(words[3].startswith("ARRAY [")) and not(words[1] == "VxDummy") ):
                                    variableNameFoundAndWritten += 1
                                    oneRow = {}                           
                                    oneRow['startOfLine']='    <LogItem'                            
                                    oneRow['ID']= str(self.IdNo)                                                                                     
                                    #oneRow['Titel'] = words[1] + " | " + variableName + " | " + variableText
                                    if (indexSubword >= 1):
                                        oneRow['Titel'] = (words[1].ljust(40)) + variableText
                                    else:
                                        oneRow['Titel'] = (words[1].ljust(40)) + variableText                     
                                    oneRow['Name'] = words[1]                              
                                    oneRow['POU'] = ''
                                    oneRow['Einheit'] = words[3]                            
                                    oneRow['Faktor'] = '1'                            
                                    oneRow['Offset'] =  '0'                            
                                    oneRow['EndOfLine'] = '></LogItem>\n'                            
                                    line = ''
                                    for rowKey in oneRow.keys():   #prepare one row of the RailMonitor conf xml file
                                        if (rowKey == 'startOfLine') or (rowKey == 'EndOfLine'):
                                            line += oneRow[rowKey] 
                                        else:
                                            line += ' ' + rowKey + '="' + oneRow[rowKey] + '"' 
                                    self.xmlFileHnd.write(line)                            
                                    self.IdNo += 1
                if not(self.varListFileHnd.closed):
                    self.varListFileHnd.close()
        print("variable named " + variableName + " found " + str(variableNameFound) + " times.")
        print("variable named " + variableName + " found and written " + str(variableNameFoundAndWritten) + " times.")

    def closeFile(self):

        if not(self.xmlFileHnd.closed):
            self.xmlFileHnd.write('  </LogResource>\n')
            self.xmlFileHnd.write('</RailMonitor>\n')
            self.xmlFileHnd.close()


    def writeLineFromPeriphSigList(self, variableName = 'VariableNameTBD', variableText = 'VariableTextTBD', varListFileName = 'nil'):
        variableNameFound = 0
        variableNameFoundAndWritten = 0
        if (varListFileName == 'nil'):
            varListFileName = self.varListFileName

        if not(self.xmlFileHnd.closed):
            if (varListFileName == 'nil'):
                self.xmlFileHnd.write('    <LogItem ID="" Titel="'  + variableText + '" Name="' + variableName + '" POU="" Einheit="" Faktor="1" Offset="0"></LogItem>\n')
            else:
                #self.varListFileHnd = open(varListFileName,'r',encoding="ansi")
                with open(varListFileName,'r',encoding="ansi") as self.varListFileHnd: #open VARLIST.CSV file 
                    words = ['']                
                    for varListLine in self.varListFileHnd:                       
                        if (-1 != varListLine.find(variableName)): #search for the variable in the VARLIST.CSV
                            variableNameFound += 1
                            words = varListLine.split(";") #split the row of the VARLIST.CSV file by the ;
                            if (((words[0].find('VAR_GLOBAL') != -1) and (words[1].find(variableName) != -1)) and (words[2].startswith("#") or words[2].startswith("%")) and not(words[1].startswith("FB_")) and not(words[3].startswith("ARRAY [")) and not(words[1] == "VxDummy")):
                                variableNameFoundAndWritten += 1
                                oneRow = {}                           
                                oneRow['startOfLine']='    <LogItem'                            
                                oneRow['ID']= str(self.IdNo)                    
                                oneRow['Titel'] = variableText
                                #oneRow['Titel'] = words[1] + " " + variableText
                                #oneRow['Titel'] = words[1]                            
                                oneRow['Name'] = words[1]                              
                                oneRow['POU'] = ''
                                oneRow['Einheit'] = words[3]                            
                                oneRow['Faktor'] = '1'                            
                                oneRow['Offset'] =  '0'                            
                                oneRow['EndOfLine'] = '></LogItem>\n'                            
                                line = ''
                                for rowKey in oneRow.keys():   #prepare one row of the RailMonitor conf xml file
                                    if (rowKey == 'startOfLine') or (rowKey == 'EndOfLine'):
                                        line += oneRow[rowKey] 
                                    else:
                                        line += ' ' + rowKey + '="' + oneRow[rowKey] + '"' 
                                self.xmlFileHnd.write(line)                            
                                self.IdNo += 1
                if not(self.varListFileHnd.closed):
                    self.varListFileHnd.close()
        print("variable named " + variableName + " found " + str(variableNameFound) + " times.")
        print("variable named " + variableName + " found and written " + str(variableNameFoundAndWritten) + " times.")
        if (variableNameFound > 0 and variableNameFoundAndWritten == 0):
            print("variable not written: " + variableName)



if __name__ == '__main__': #argv[1] - conversion file converting name of the variable to the defined text, example of one line of the file: CL_F=DoorsLogic
                           #argv[2] - name of the VARLIST.CSV file generated by the CAP1131 tool
                           #argv[3] - name of the RailMonitor configuration xml file, created by this python script 
                           #argv[4] - name of the excel file "Periphery signal list"
    #try:
        doGlobalVars = True # parameter defining if the xml files based on the var list shall be generated
        print (len(sys.argv))
        line = ''
        words = ['', '']
        wordOld = ''
        parametr = { }
        parametrList = []
        index = 0
        nameOfFile = []
        sheetList = []

        if ((len(sys.argv) > 3) and (doGlobalVars == True)):
            NameConversion = ClNameConversion(sys.argv[1]) #open a conversion file translating name of the global vars to the comment text
            XmlFile1 = ClXmlFile(sys.argv[3])              #create a new xml file named as stated in the argv[3]

            for file in os.listdir("."):
                if file.endswith(".rm"):  
                    with open(file,'r',encoding="utf8") as inpFile:     # read input file listOfVariables  
                        i = 0
                        parametrList = []
                        for line in inpFile:                
                            if (-1 != line.find("<LogResource") and isinstance(XmlFile1, ClXmlFile)): #create a new xml file when the <LogResource tag has been read frome the input file
                                 XmlFile1.closeFile()
                                 del XmlFile1
                                 index += 1 
                                 nameOfFile = line.split("=\"")
                                 nameOfFile1 = nameOfFile[1].split("\"") 
                                 XmlFile1 = ClXmlFile(nameOfFile1[0] + ".xml", sys.argv[2], line)     
                                 subwordIndex = 1
                            #dirtyWords = line.split(' - ')   
                            dirtyWords = line.split(': ')                                  
                            if (len(dirtyWords) > 1):
                                words[0] = dirtyWords[0].strip()
                                words[1] = dirtyWords[1].strip()
                                if ((wordOld != words[0]) and file == 'globalvars.rm'):  #create new xml file when new global var has been read from the ctrl+C/ctrl+V copy of the global vars
                                    if (isinstance(XmlFile1, ClXmlFile)):
                                        XmlFile1.closeFile()
                                        del XmlFile1
                                        index += 1 
                                        tabname = NameConversion.searchVars(words[0])
                                        tabname = words[0] + "-" + tabname
                                        logResource = "<LogResource Name=\"" + str(tabname) + "\" Bild=\"pictures\\trampng.png\">\n"
                                       
                                       
                                        #XmlFile1 = ClXmlFile(words[0] + ".xml", sys.argv[2], logResource)
                                        XmlFile1 = ClXmlFile(tabname + ".xml", sys.argv[2], logResource)
                                        subwordIndex = 0

                                wordOld = words[0]


                                parametr = {'name': words[0], 'text': words[1]}
                                parametrList.append(parametr) 
                                XmlFile1.writeLine(parametrList[i]['name'], parametrList[i]['text'][0:60], sys.argv[2], subwordIndex)
                                i += 1
                                if i > 200: #to skip form the input file after first number of lines
                                    break
            XmlFile1.closeFile()
            if not(inpFile.closed):
                inpFile.close()

        if (len(sys.argv) > 4):
            excelFile = sys.argv[4]
            #sheet = pe.get_sheet(excelFile)

            #wb = openpyxl.load_workbook(excelFile)     # read input file PR_2163954.xlsm
            wb = openpyxl.load_workbook(filename=excelFile, read_only=True)
            sheetNameList = ['114', '147A', '165A', '166A', '182A', '183A', '211L', '256A']
            for sheetName in sheetNameList:
                #sheetName = '114'

                ws = wb[sheetName]
            
                index += 1 
                logResource = "<LogResource Name=\"" + sheetName + "\" Bild=\"pictures\\trampng.png\">\n"
                XmlFile1 = ClXmlFile(sheetName + ".xml", sys.argv[2], logResource)


                for row in ws.rows:
                    if (row[0].value == '-'):
                        if (row[0].row >= 2):
                            if ((row[7].value != None) and (row[8].value != None)):
                                text = row[1].value.ljust(7) + " " + row[3].value.ljust(19) + " " + row[4].value.ljust(3) + " " + row[5].value.ljust(3) + " " + row[7].value.ljust(35) + " " + row[8].value.ljust(35) 
                                XmlFile1.writeLineFromPeriphSigList(row[7].value, text, sys.argv[2])

                XmlFile1.closeFile()
                    
            """
                                            for cell in row:
                                                if (cell.value != None):
                                                    print(cell.value)
                                        cell1 = ws['A1']
                                        print(cell1.value)
                                        cell1 = ws['C10']
                                        print(cell1.value)
                                        cell1 = ws['H13']
                                        print(cell1.value)
                                        pass
            """

       
            






    #except IOError:
        #print("Error: can't open, read or write to/from file")
    #except Exception:
        #print( "Error: the Exception occured.")